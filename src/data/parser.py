"""
parser.py – Excel workbook ingestion for Siteworks.

PARSING ASSUMPTIONS (documented here per project spec):
========================================================
1. Preferred: multi-sheet workbook layout:
     • **Scores** — wide table (row 1 = headers; City, Name; then sub-account scores).
     • **City Assignments** — optional City ↔ numeric code (optional if Names are in Scores).
     • **Values** — optional same layout as Scores with raw measurements → raw_value.
     • **Account Weights** — optional Account | Weight for category defaults (must match CATEGORIES).
     • **Sources** — optional City | Name … wide URLs (same columns as **Scores**) for per-city citations.
2. Legacy single-sheet layouts if multi-sheet parse does not yield ≥3 cities:
   a) Sheet "Site Selector Data - edited" (or candidates in _find_sheet):
      Wide matrix OR long / tidy (row 5 headers; rows 2–3 city codes).
3. Score cells contain numeric values in the 1–5 range (see _parse_score).
4. Columns that do not map to schema subcategories (e.g. Seismic) are skipped.

If the workbook is absent or unreadable the loader falls back to the
built-in pilot dataset defined in loader.py.
"""

import logging
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from src.data.schema import (
    CATEGORIES,
    SUBCATEGORIES,
    PILOT_CITIES,
    SubcategoryScore,
    CityData,
)
from src.logic.scoring import normalize_weights

logger = logging.getLogger(__name__)

# Workbook column titles (Scores / Values) → canonical SUBCATEGORIES keys.
# Columns with no mapping (e.g. Seismic) are skipped for scoring.
_HEADER_ALIASES: Dict[str, str] = {
    "baseline water stress (regional)": "Baseline Water Stress",
    "annual precipitation (1991-2020)": "Annual Precipitation",
    "recycled water infrastructure": "Recycled Water Infrastructure",
    "cooling degree days (1991-2020)": "Cooling Degree Days",
    "annual mean relative humidity": "Annual Mean Humidity",
    "carbon regulations": "Grid Carbon Intensity",
    "renewable energy mix": "Renewable Energy Mix",
    "industrial electricity rate (2024)": "Industrial Electricity Rate",
    "water & sewer cost (industrial)": "Water & Sewer Cost",
    "environmental justice index": "Environmental Justice Index",
    "seismic hazard (usgs 2023)": "Seismic Hazard",
    "flood risk zone": "Flood Risk",
    "tornado frequency (annual avg)": "Tornado Frequency",
    "wildfire hazard": "Wildlife Hazard",
    "winter weather disruption": "Winter Weather Disruption",
    "protected area proximity": "Protected Area Proximity",
}

# ---------------------------------------------------------------------------
# Public parse entry point
# ---------------------------------------------------------------------------

def parse_rh_workbook(
    path: Path,
) -> Tuple[
    Optional[Dict[str, CityData]],
    Optional[Dict[str, float]],
    Optional[Dict[str, List[Tuple[str, str]]]],
]:
    """
    Parse `Data_Center_Site_Selector_RH.xlsx`.

    Returns:
        (city_data_map or None, category_weights or None, sources_by_subcategory or None)
        – **sources_by_subcategory** maps each subcategory key to [(city_name, url), …]
          from the **Sources** sheet when present; else ``None``.

    Returns ``(None, None, None)`` if openpyxl is unavailable or the file cannot be read.
    """
    try:
        import openpyxl  # noqa: PLC0415  (optional dependency)
    except ImportError:
        logger.warning("openpyxl not installed – cannot read workbook.")
        return None, None, None

    if not path.is_file():
        logger.info("Workbook not found at %s.", path)
        return None, None, None

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.error("Failed to open workbook %s: %s", path, exc)
        return None, None, None

    try:
        sources_map = _parse_sources_sheet(wb)
        parsed_ms, weights_ms = _try_parse_multisheet_workbook(wb)
        if parsed_ms and len(parsed_ms) >= 1:
            return parsed_ms, weights_ms, sources_map

        sheet = _find_sheet(wb)
        if sheet is None:
            logger.warning("Could not locate a data sheet for legacy parsing.")
            return None, None, sources_map

        if _sheet_looks_long_format(sheet):
            parsed = _extract_scores_long(sheet)
            if parsed:
                return parsed, None, sources_map
            logger.error(
                "Sheet appears to be long format, but long-format parsing returned no data. "
                "Refusing fallback to wide parser to avoid misinterpreting workbook."
            )
            return {}, None, sources_map

        legacy = _extract_scores(sheet)
        return legacy if legacy else None, None, sources_map
    finally:
        wb.close()


def _try_parse_multisheet_workbook(wb) -> Tuple[Optional[Dict[str, CityData]], Optional[Dict[str, float]]]:
    """Parse **Scores** (+ optional **Values**, **Account Weights**, **City Assignments**)."""
    if "Scores" not in wb.sheetnames:
        return None, None

    city_by_code = {}
    if "City Assignments" in wb.sheetnames:
        city_by_code = _parse_city_assignments_sheet(wb["City Assignments"])

    parsed = _parse_scores_wide_sheet(wb["Scores"], city_by_code)
    if not parsed:
        return None, None

    if "Values" in wb.sheetnames:
        _merge_values_wide_sheet(wb["Values"], parsed)

    weights: Optional[Dict[str, float]] = None
    if "Account Weights" in wb.sheetnames:
        weights = _parse_account_weights_sheet(wb["Account Weights"])

    return parsed, weights


def _header_to_canonical(header: str, all_subs: Dict[str, str]) -> Optional[str]:
    low = header.lower().strip()
    if low in _HEADER_ALIASES:
        return _HEADER_ALIASES[low]
    if "seismic" in low and "hazard" in low:
        return "Seismic Hazard"
    if "wildfire" in low:
        return "Wildlife Hazard"
    norm = _normalize_workbook_sub_label(header)
    found = _match_subcategory(norm, all_subs)
    if found:
        return found
    return _match_subcategory(header, all_subs)


def _parse_city_assignments_sheet(sheet) -> Dict[int, str]:
    """**City Assignments** sheet: City (col A), Code (col B)."""
    out: Dict[int, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        code_raw = row[1] if len(row) > 1 else None
        try:
            code = int(float(str(code_raw).strip()))
        except (TypeError, ValueError):
            continue
        if name:
            out[code] = name
    return out


def _extract_url_from_cell(raw) -> Optional[str]:
    """First HTTP(S) URL from a cell; strip trailing notes after a space + opening parenthesis."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s.lower().startswith("http"):
        return None
    for sep in (" (", "\n", "\t"):
        if sep in s:
            s = s.split(sep, 1)[0].strip()
    while s.endswith((".", ",")):
        s = s[:-1]
    return s or None


def _parse_sources_sheet(wb) -> Optional[Dict[str, List[Tuple[str, str]]]]:
    """**Sources**: same wide layout as **Scores** (Name + per-column URLs)."""
    if "Sources" not in wb.sheetnames:
        return None
    sheet = wb["Sources"]
    row1 = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    if not row1 or not row1[0]:
        return {}

    headers = [
        str(h).strip() if h is not None else "" for h in row1[0]
    ]
    if "Name" not in headers:
        logger.warning("Sources sheet: no **Name** column — skipping citation parse.")
        return {}

    idx_name = headers.index("Name")
    all_subs = {
        sub: sub for subs in SUBCATEGORIES.values() for sub in subs
    }
    col_canon: List[Tuple[int, str]] = []
    for i in range(2, len(headers)):
        h = headers[i]
        if not h:
            continue
        canon = _header_to_canonical(h, all_subs)
        if canon:
            col_canon.append((i, canon))

    if not col_canon:
        return {}

    by_sub: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= idx_name or row[idx_name] is None:
            continue
        city_name = str(row[idx_name]).strip()
        if not city_name:
            continue
        for idx, canon in col_canon:
            raw = row[idx] if len(row) > idx else None
            url = _extract_url_from_cell(raw)
            if url:
                by_sub[canon].append((city_name, url))

    return dict(by_sub)


def _parse_scores_wide_sheet(
    sheet,
    city_by_code: Dict[int, str],
) -> Optional[Dict[str, CityData]]:
    """**Scores** sheet: row 1 headers; col A = city code; col B = city name; scores from col C."""
    row1 = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    if not row1:
        return None
    headers = row1[0]
    all_subs = {
        sub: sub for subs in SUBCATEGORIES.values() for sub in subs
    }
    col_canon: List[Tuple[int, str]] = []
    for i in range(2, len(headers)):
        h = headers[i]
        if h is None or not str(h).strip():
            continue
        canon = _header_to_canonical(str(h).strip(), all_subs)
        if canon:
            col_canon.append((i, canon))
        else:
            logger.info("Scores sheet: skipping unmapped column %r", h)

    city_data: Dict[str, CityData] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or row[1] is None:
            continue
        city_name = str(row[1]).strip()
        if not city_name:
            continue

        if city_by_code and row[0] is not None:
            try:
                code = int(float(str(row[0]).strip()))
                mapped = city_by_code.get(code)
                if mapped and mapped.lower() != city_name.lower():
                    logger.warning(
                        "City code %s maps to %r but row Name=%r — using Name column.",
                        code,
                        mapped,
                        city_name,
                    )
            except (TypeError, ValueError):
                pass

        if city_name not in city_data:
            city_data[city_name] = CityData(name=city_name)

        for idx, canon in col_canon:
            raw_score = row[idx] if len(row) > idx else None
            score = _parse_score(raw_score)
            city_data[city_name].subcategory_scores[canon] = SubcategoryScore(
                name=canon,
                score=score if score is not None else float("nan"),
                raw_value=None,
                note=(
                    "Parsed from workbook (Scores)"
                    if score is not None
                    else "Missing/blank"
                ),
            )
            if score is None:
                city_data[city_name].data_quality_notes.append(
                    f"Missing score for '{canon}'"
                )

    return city_data if city_data else None


def _merge_values_wide_sheet(values_sheet, city_data: Dict[str, CityData]) -> None:
    """Merge **Values** sheet raw measurements into existing SubcategoryScore.raw_value."""
    row1 = next(values_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    all_subs = {
        sub: sub for subs in SUBCATEGORIES.values() for sub in subs
    }
    col_canon: List[Tuple[int, str]] = []
    for i in range(2, len(row1)):
        h = row1[i]
        if h is None or not str(h).strip():
            continue
        canon = _header_to_canonical(str(h).strip(), all_subs)
        if canon:
            col_canon.append((i, canon))

    for row in values_sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2 or row[1] is None:
            continue
        city_name = str(row[1]).strip()
        if city_name not in city_data:
            continue
        cd = city_data[city_name]
        for idx, canon in col_canon:
            raw = row[idx] if len(row) > idx else None
            if canon not in cd.subcategory_scores:
                continue
            entry = cd.subcategory_scores[canon]
            if raw is not None:
                entry.raw_value = str(raw).strip()


def _parse_account_weights_sheet(sheet) -> Optional[Dict[str, float]]:
    """**Account Weights**: Account | Weight for each entry in CATEGORIES."""
    raw: Dict[str, float] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        try:
            f = float(row[1])
        except (TypeError, ValueError, IndexError):
            continue
        for cat in CATEGORIES:
            if cat.lower() == label.lower():
                raw[cat] = f
                break

    if len(raw) != len(CATEGORIES):
        logger.warning(
            "Account Weights sheet: need %s categories, found %s keys.",
            len(CATEGORIES),
            len(raw),
        )
        return None
    try:
        return normalize_weights(raw)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Internal helpers (legacy single-sheet layouts)
# ---------------------------------------------------------------------------

def _find_sheet(wb):
    """Return the target worksheet, trying several likely names."""
    candidates = [
        "Site Selector Data - edited",
        "Site Selector Data",
        "SiteSelector",
        "Data",
        "Sheet1",
    ]
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name]
    # Fall back to first sheet
    if wb.sheetnames:
        logger.warning(
            "Target sheet not found; falling back to first sheet: %s",
            wb.sheetnames[0],
        )
        return wb[wb.sheetnames[0]]
    return None


def _sheet_looks_long_format(sheet) -> bool:
    """True if row 5 looks like the tidy layout (Sub-Account, City, Score columns)."""
    row5_rows = list(sheet.iter_rows(min_row=5, max_row=5, values_only=True))
    if not row5_rows:
        return False
    row5 = row5_rows[0]
    if not row5 or len(row5) < 7:
        return False
    joined = " ".join(
        str(x).strip().lower() for x in row5[:8] if x is not None and str(x).strip()
    )
    return "sub-account" in joined and "city" in joined and "score" in joined


def _city_nums_from_assignment_header(sheet) -> Dict[int, str]:
    """
    Build city index → display name from rows 2–3 (numbered columns under
    'City Assignments' in the RH workbook).
    """
    r2_rows = list(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    r3_rows = list(sheet.iter_rows(min_row=3, max_row=3, values_only=True))
    if not r2_rows or not r3_rows:
        return {}
    r2 = r2_rows[0]
    r3 = r3_rows[0]
    out: Dict[int, str] = {}
    for i, h in enumerate(r2):
        if h is None:
            continue
        try:
            num = int(float(str(h).strip()))
        except (TypeError, ValueError):
            continue
        if num < 1:
            continue
        if i >= len(r3) or r3[i] is None:
            continue
        name = str(r3[i]).strip()
        if name:
            out[num] = name
    return out


def _normalize_workbook_sub_label(label: str) -> str:
    """Map workbook Sub-Account wording to strings that match SUBCATEGORIES."""
    low = label.lower()
    if "wildfire hazard" in low:
        return "Wildlife Hazard"
    if "seismic" in low and "hazard" in low:
        return "Seismic Hazard"
    return label


def _extract_scores_long(sheet) -> Dict[str, CityData]:
    """
    Parse the long / tidy layout used in the distributed RH workbook:
    header on row 5; blocks of rows with Sub-Account in column B, City index
    in column E, numeric score in column G, raw measurement in column F.
    """
    city_by_num = _city_nums_from_assignment_header(sheet)
    if len(city_by_num) < 1:
        logger.warning("Long-format sheet: no city index → name mapping in rows 2–3.")
        return {}

    all_subs = {
        sub: sub
        for subs in SUBCATEGORIES.values()
        for sub in subs
    }

    city_data: Dict[str, CityData] = {
        name: CityData(name=name) for name in city_by_num.values()
    }

    current_sub_label: Optional[str] = None
    for row in sheet.iter_rows(min_row=6, values_only=True):
        if not row:
            continue
        b = row[1]
        if b is not None and str(b).strip():
            current_sub_label = str(b).strip()

        if not current_sub_label:
            continue

        city_key = row[4] if len(row) > 4 else None
        raw_measure = row[5] if len(row) > 5 else None
        score_raw = row[6] if len(row) > 6 else None

        try:
            city_num = int(float(city_key)) if city_key is not None else None
        except (TypeError, ValueError):
            continue
        city_name = city_by_num.get(city_num) if city_num is not None else None
        if not city_name:
            continue

        norm_label = _normalize_workbook_sub_label(current_sub_label)
        canonical = _match_subcategory(norm_label, all_subs)
        if canonical is None:
            canonical = _match_subcategory(current_sub_label, all_subs)
        if canonical is None:
            continue

        score = _parse_score(score_raw)
        city_data[city_name].subcategory_scores[canonical] = SubcategoryScore(
            name=canonical,
            score=score if score is not None else float("nan"),
            raw_value=str(raw_measure) if raw_measure is not None else None,
            note="Parsed from workbook" if score is not None else "Missing/blank",
        )
        if score is None:
            city_data[city_name].data_quality_notes.append(
                f"Missing score for '{canonical}'"
            )

    return city_data


def _extract_scores(sheet) -> Dict[str, CityData]:
    """
    Extract city × subcategory score matrix from the worksheet.

    Expected layout:
        Row 1 : headers – col A = label, col B+ = city names
        Row 2+: data   – col A = subcategory name, col B+ = score values
    """
    # Read header row to find city columns
    header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    city_col_map: Dict[str, int] = {}  # city_name → 0-based col index
    for col_idx, cell_val in enumerate(header_row):
        if col_idx == 0:
            continue
        if cell_val:
            cell_str = str(cell_val).strip()
            # Match against known pilot city names (case-insensitive)
            matched = _match_city(cell_str)
            if matched:
                city_col_map[matched] = col_idx

    if not city_col_map:
        logger.warning("No pilot cities found in workbook header row.")
        return {}

    # Build subcategory name → canonical name mapping
    all_subs = {
        sub: sub
        for subs in SUBCATEGORIES.values()
        for sub in subs
    }

    # Initialise CityData containers
    city_data: Dict[str, CityData] = {
        city: CityData(name=city) for city in city_col_map
    }

    # Read data rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        row_label = str(row[0]).strip()
        canonical = _match_subcategory(row_label, all_subs)
        if canonical is None:
            continue

        for city, col_idx in city_col_map.items():
            raw_val = row[col_idx] if col_idx < len(row) else None
            score = _parse_score(raw_val)
            city_data[city].subcategory_scores[canonical] = SubcategoryScore(
                name=canonical,
                score=score if score is not None else float("nan"),
                raw_value=str(raw_val) if raw_val is not None else None,
                note="Parsed from workbook" if score is not None else "Missing/blank",
            )
            if score is None:
                city_data[city].data_quality_notes.append(
                    f"Missing score for '{canonical}'"
                )

    return city_data


def _match_city(cell_str: str) -> Optional[str]:
    """Fuzzy-match a cell string to a known pilot city name."""
    cell_lower = cell_str.lower()
    for city in PILOT_CITIES:
        if city.lower() in cell_lower or cell_lower in city.lower():
            return city
    return None


def _match_subcategory(label: str, mapping: Dict[str, str]) -> Optional[str]:
    """Fuzzy-match a row label to a canonical subcategory name."""
    label_lower = label.lower().strip()
    for canonical in mapping:
        if canonical.lower() == label_lower:
            return canonical
        # Allow partial match (at least 70% of words)
        canonical_words = set(canonical.lower().split())
        label_words = set(label_lower.split())
        if canonical_words and label_words:
            overlap = canonical_words & label_words
            if len(overlap) / len(canonical_words) >= 0.7:
                return canonical
    return None


def _parse_score(value) -> Optional[float]:
    """Convert a cell value to a float score, or None if invalid."""
    if value is None:
        return None
    try:
        f = float(value)
        if 0 <= f <= 5:
            return f
        # Some workbooks store raw ranks (1=best) – invert if > 5
        if 1 <= f <= 10:
            return round(6 - (f / 2), 2)  # normalise to 1–5
        return None
    except (TypeError, ValueError):
        return None
